import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook()
ws = wb.active
ws.title = "แปลงเงิน"

ws['A1'] = "แปลงค่าเงิน ดอลลาร์ → บาทไทย"
ws['A1'].font = Font(size=16, bold=True)
ws.merge_cells('A1:B1')

ws['A3'] = "ค่าเงินดอลลาร์ (USD):"
ws['B3'] = 0
ws['B3'].number_format = '#,##0.00'

ws['A4'] = "อัตราแลกเปลี่ยน (บาท/USD):"
ws['B4'] = 35.50
ws['B4'].number_format = '#,##0.00'

ws['A5'] = "ผลลัพธ์ (บาท):"
ws['B5'] = "=B3*B4"
ws['B5'].number_format = '#,##0.00'
ws['A5'].font = Font(bold=True)
ws['B5'].font = Font(bold=True, color="FF0000")

ws['A3'].alignment = Alignment(horizontal='right')
ws['A4'].alignment = Alignment(horizontal='right')
ws['A5'].alignment = Alignment(horizontal='right')

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18

ws['A7'] = "📌 วิธีใช้:"
ws['A7'].font = Font(bold=True, size=12)
ws['A8'] = "1. ใส่ค่าเงินดอลลาร์ในช่อง B3"
ws['A9'] = "2. แก้ไขอัตราแลกเปลี่ยนในช่อง B4 (ถ้าต้องการ)"
ws['A10'] = "3. กดปุ่มคำนวณ หรือ แก้ไขค่าแล้วผลจะอัพเดทอัตโนมัติ"

for row in range(8, 11):
    ws[f'A{row}'].font = Font(size=10)

ws.column_dimensions['A'].width = 50

wb.save("แปลงเงิน.xlsm")

print("=" * 50)
print("✅ สร้างไฟล์ Excel สำเร็จ: แปลงเงิน.xlsm")
print("=" * 50)
print("\n📋 วิธีเพิ่มปุ่มกด (VBA Macro):")
print("-" * 50)
print("1. เปิดไฟล์ แปลงเงิน.xlsm ใน Excel")
print("2. กด Alt+F11 เพื่อเปิด VBA Editor")
print("3. คลิกขวา → Insert → Module")
print("4. วางโค้ดนี้:\n")

vba_code = """Sub ConvertCurrency()
    Dim usdValue As Double
    Dim rate As Double
    Dim result As Double
    
    usdValue = Range("B3").Value
    rate = Range("B4").Value
    
    result = usdValue * rate
    
    Range("B5").Value = result
    
    MsgBox usdValue & " USD = " & Format(result, "#,##0.00") & " บาท", _
           vbInformation, "แปลงเงินสำเร็จ"
End Sub
"""
print(vba_code)
print("-" * 50)
print("5. บันทึกและกลับไปที่ Excel")
print("6. กด Alt+F8 เลือก ConvertCurrency แล้วกด Run")
print("\n💡 หมายเหตุ: สูตรใน B5 จะคำนวณอัตโนมัติเมื่อเปลี่ยนค่า")
