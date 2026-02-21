import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Currency"

ws['A1'] = "Convert USD to THB"
ws['A1'].font = Font(size=16, bold=True)
ws.merge_cells('A1:B1')

ws['A3'] = "USD Amount:"
ws['B3'] = 100
ws['B3'].number_format = '#,##0.00'

ws['A4'] = "Exchange Rate (THB/USD):"
ws['B4'] = 35.50
ws['B4'].number_format = '#,##0.00'

ws['A5'] = "Result (THB):"
ws['B5'] = "=B3*B4"
ws['B5'].number_format = '#,##0.00'
ws['A5'].font = Font(bold=True)
ws['B5'].font = Font(bold=True, color="FF0000")

ws['A3'].alignment = Alignment(horizontal='right')
ws['A4'].alignment = Alignment(horizontal='right')
ws['A5'].alignment = Alignment(horizontal='right')

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18

ws['A7'] = "INSTRUCTIONS:"
ws['A7'].font = Font(bold=True, size=12)
ws['A8'] = "1. Enter USD amount in cell B3"
ws['A9'] = "2. Change exchange rate in cell B4 (optional)"
ws['A10'] = "3. Result in B5 calculates automatically"
ws['A11'] = ""
ws['A12'] = "To add VBA button:"
ws['A13'] = "1. Save as .xlsm manually"
ws['A14'] = "2. Press Alt+F11 in Excel"
ws['A15'] = "3. Insert > Module"
ws['A16'] = "4. Paste the VBA code"

ws.column_dimensions['A'].width = 45

wb.save("convert_usd.xlsx")
print("Excel file created: convert_usd.xlsx")
