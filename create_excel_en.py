import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Currency"

ws['A1'] = "Convert USD to THB"
ws['A1'].font = Font(size=16, bold=True)
ws.merge_cells('A1:B1')

ws['A3'] = "USD Amount:"
ws['B3'] = 0
ws['B3'].number_format = '#,##0.00'

ws['A4'] = "Exchange Rate (THB/USD):"
ws['B4'] = 35.50
ws['B4'].number_format = '#,##0.00'

ws['A5'] = "Result (THB):"
ws['B5'] = "=B3*B4"
ws['B5'].number_format = '#,##0.00'
ws['A5'].font = Font(bold=True)
ws['B5'].font = Font(bold=True, color="FF0000")

ws['A3'].alignment = Alignment(horizontal='right')
ws['A4'].alignment = Alignment(horizontal='right')
ws['A5'].alignment = Alignment(horizontal='right')

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18

ws['A7'] = "Instructions:"
ws['A7'].font = Font(bold=True, size=12)
ws['A8'] = "1. Enter USD amount in B3"
ws['A9'] = "2. Change exchange rate in B4 (optional)"
ws['A10'] = "3. Result in B5 updates automatically"

ws.column_dimensions['A'].width = 40

wb.save("convert_usd_to_thb.xlsm")

print("Excel file created: convert_usd_to_thb.xlsm")
print("\nHow to add VBA button:")
print("1. Open convert_usd_to_thb.xlsm in Excel")
print("2. Press Alt+F11 to open VBA Editor")
print("3. Right-click → Insert → Module")
print("4. Paste this code:\n")

vba_code = """Sub ConvertCurrency()
    Dim usdValue As Double
    Dim rate As Double
    Dim result As Double
    
    usdValue = Range("B3").Value
    rate = Range("B4").Value
    
    result = usdValue * rate
    
    Range("B5").Value = result
    
    MsgBox usdValue & " USD = " & Format(result, "#,##0.00") & " THB", _
           vbInformation, "Conversion Complete"
End Sub
"""
print(vba_code)
print("5. Save and press Alt+F8 to run ConvertCurrency")
